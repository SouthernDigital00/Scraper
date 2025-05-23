import asyncio
import openpyxl
import re
from urllib.parse import urlparse
from playwright.async_api import async_playwright

# Configuration
EXCEL_FILE = r"C:\Users\ADickerson.RISEMFG\OneDrive - Rise Manufacturing\Desktop\Work Creations\CustomerInventory_Extract_Data.xlsx"
SHEET_NAME = 'Inventory Scrape'

async def dismiss_popups_primaryarms(page):
    # Age gate
    for sel in ["button:has-text('I am 18 or older')", "button:has-text('Yes, I am 18 or older')", "button:has-text('Yes')"]:
        try:
            await page.wait_for_selector(sel, timeout=3000)
            await page.click(sel)
            await page.wait_for_timeout(500)
            break
        except:
            pass
    # Cookie banner
    for sel in ["button:has-text('Accept & Close')", "button:has-text('Accept')", "button:has-text('I Accept')"]:
        try:
            await page.wait_for_selector(sel, timeout=3000)
            await page.click(sel)
            await page.wait_for_timeout(500)
            break
        except:
            pass
    # Email signup popup
    try:
        btn = page.locator("button:has-text('No thanks')")
        if await btn.is_visible():
            await btn.click()
            await page.wait_for_timeout(500)
    except:
        pass
    # Remove any remaining dialogs
    try:
        await page.evaluate("() => document.querySelectorAll('[role=dialog]').forEach(el=>el.remove())")
    except:
        pass

async def extract_price_and_stock_ar15(page, shape, color):
    # Select shape and color
    await page.click(f"button[data-shape='{shape}']", timeout=5000)
    await page.click(f"button[data-color='{color}']", timeout=5000)
    # Add to cart
    await page.click("button:has-text('Add to cart')", timeout=5000)
    # Wait for mini-cart price
    await page.wait_for_selector('.woocommerce-mini-cart .woocommerce-Price-amount', timeout=5000)
    text = (await page.text_content('.woocommerce-mini-cart .woocommerce-Price-amount')).strip()
    price = float(re.search(r'\d+(?:\.\d+)?', text).group())
    # Read inventory max
    max_inv = int(await page.get_attribute('input.qty', 'max') or 0)
    return price, max_inv

async def extract_price_and_stock_primaryarms(page):
    # Dismiss pop-ups
    await dismiss_popups_primaryarms(page)
    # Click add-to-cart
    await page.click('button[data-action="sticky"][data-type="add-to-cart"]', timeout=5000)
    await page.wait_for_selector('.cart-confirmation-modal', timeout=5000)
    # Parse price
    price_text = (await page.text_content('.cart-confirmation-modal span.transaction-line-views-price-exact')).strip()
    price = float(re.search(r'\d+(?:\.\d+)?', price_text).group())
    # Try availability link
    avail = await page.get_attribute('link[itemprop="availability"]', 'href') or ''
    if 'InStock' not in avail:
        inventory = 0
    else:
        # Fallback: increment product page qty to determine max
        count = 1
        while True:
            try:
                await page.click("button.product-details-quantity-add", timeout=1000)
                count += 1
                await page.wait_for_timeout(200)
            except:
                break
        inventory = count
    return price, inventory

async def run():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        page = await context.new_page()

        for row in ws.iter_rows(min_row=2, values_only=False):
            sku = row[1].value
            url = row[2].value
            print(f"🔍 {sku}")
            if not url or not url.startswith('http'):
                row[3].value = 'Invalid URL'
                continue

            await page.goto(url, timeout=30000)
            await page.wait_for_timeout(2000)
            domain = urlparse(url).netloc.lower()

            try:
                if 'ar15discounts.com' in domain:
                    price, inv = await extract_price_and_stock_ar15(page, 'Flat', 'Black')
                elif 'primaryarms.com' in domain:
                    price, inv = await extract_price_and_stock_primaryarms(page)
                else:
                    price, inv = 'Error', 0

                if inv == 0 or price == 'Error':
                    row[3].value = 'n/a'
                    row[4].value = 'Out of Stock'
                    print(f"❌ {sku} → Out of Stock")
                else:
                    row[3].value = price
                    row[4].value = inv
                    print(f"✅ {sku} → Price: {price} | Inventory: {inv}")
            except Exception as e:
                row[3].value = 'n/a'
                row[4].value = 'Out of Stock'
                print(f"❌ {sku} flow failed: {e}")

        await browser.close()
        wb.save(EXCEL_FILE)
        print("✅ Excel updated.")

if __name__ == '__main__':
    asyncio.run(run())
